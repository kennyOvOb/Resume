import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Color, Side, Border
from pathlib import Path
import time


class SummaryStyle:
    def __init__(self, summary_path):
        self.summary_path = summary_path
        self.sheet1_name = "單客戶"
        self.sheet2_name = "多客戶"
        self.sheet_list = [self.sheet1_name, self.sheet2_name]
        self.one_site_columns = {
            "客戶": "A",
            "預設": "B",
            "序号": "C",
            "日期": "D",
            "憑證種類": "E",
            "人数": "F",
            "笔数": "G",
            "金额": "H",
            "币别": "I",
            "路徑": "J",
            "憑證人數": "K",
            "憑證筆數": "L",
            "憑證總額": "M",
            "人數差額": "N",
            "筆數差額": "O",
            "金額差额": "P",
            "檢查": "Q"
        }
        self.total_site_columns = {
            "客戶": "A",
            "預設": "B",
            "序号": "C",
            "日期": "D",
            "憑證種類": "E",
            "人数": "F",
            "笔数": "G",
            "金额": "H",
            "币别": "I",
            "子客戶": "J",
            "路徑": "K",
            "憑證人數": "L",
            "憑證筆數": "M",
            "憑證總額": "N",
            "人數差額": "O",
            "筆數差額": "P",
            "金額差额": "Q",
            "檢查": "R"
        }
        self.certificate_columns = ["憑證人數", "憑證筆數", "憑證總額"]
        self.client = 20
        self.index = 4.5703125
        self.date = 11.140625
        self.subject = 9.0
        self.member = 9
        self.quantity = 9
        self.amount = 19.42578125
        self.currency = 5.140625
        self.certificate_amount = 11.0
        self.site = 20

    # first Parameters
    @staticmethod
    def header():
        font = Font(name='楷体', size=11.0, bold=True, italic=False, color='FFFFFF')
        fill = PatternFill(fill_type="solid", fgColor="5B9BD5")
        left_border = Side(border_style='thin',
                           color=Color(rgb=None, indexed=None, auto=True, theme=None, tint=0.0, type='auto'))
        right_border = Side(border_style='thin',
                            color=Color(rgb=None, indexed=None, auto=True, theme=None, tint=0.0, type='auto'))
        top_border = Side(border_style='thick',
                          color=Color(rgb=None, indexed=None, auto=True, theme=None, tint=0.0, type='auto'))
        bottom_border = Side(border_style='thick',
                             color=Color(rgb=None, indexed=None, auto=True, theme=None, tint=0.0, type='auto'))
        border = Border(left=left_border, top=top_border, right=right_border, bottom=bottom_border)
        alignment = Alignment(horizontal="center", vertical="center")
        return font, fill, border, alignment

    def odd_row(self):  # 单数
        font = Font(name='楷体', size=11.0, bold=False, italic=False, color='00000000')
        fill = PatternFill(fill_type="solid", fgColor="FFFFFFFF")
        border = self.content_border()
        alignment = Alignment(horizontal="general", vertical="center")
        return font, fill, border, alignment

    def even_row(self):  # 双数
        font = Font(name='楷体', size=11.0, bold=False, italic=False, color='00000000')
        fill = PatternFill(fill_type="solid", fgColor="FFDDEBF7")
        border = self.content_border()
        alignment = Alignment(horizontal="general", vertical="center")
        return font, fill, border, alignment

    @staticmethod
    def content_border():
        left_border = Side(border_style='thin',
                           color=Color(rgb=None, indexed=None, auto=True, theme=None, tint=0.0, type='auto'))
        right_border = Side(border_style='thin',
                            color=Color(rgb=None, indexed=None, auto=True, theme=None, tint=0.0, type='auto'))
        top_border = Side()
        bottom_border = Side(border_style='thin',
                             color=Color(rgb=None, indexed=None, auto=True, theme=None, tint=0.0, type='auto'))
        border = Border(left=left_border, top=top_border, right=right_border, bottom=bottom_border)
        return border

    @staticmethod
    def far_left_border():
        left_border = Side(border_style='thick',
                           color=Color(rgb=None, indexed=None, auto=True, theme=None, tint=0.0, type='auto'))
        right_border = Side(border_style='thin',
                            color=Color(rgb=None, indexed=None, auto=True, theme=None, tint=0.0, type='auto'))
        top_border = Side()
        bottom_border = Side(border_style='thin',
                             color=Color(rgb=None, indexed=None, auto=True, theme=None, tint=0.0, type='auto'))
        border = Border(left=left_border, top=top_border, right=right_border, bottom=bottom_border)
        return border

    @staticmethod
    def far_right_border():
        left_border = Side(border_style='thin',
                           color=Color(rgb=None, indexed=None, auto=True, theme=None, tint=0.0, type='auto'))
        right_border = Side(border_style='thick',
                            color=Color(rgb=None, indexed=None, auto=True, theme=None, tint=0.0, type='auto'))
        top_border = Side()
        bottom_border = Side(border_style='thin',
                             color=Color(rgb=None, indexed=None, auto=True, theme=None, tint=0.0, type='auto'))
        border = Border(left=left_border, top=top_border, right=right_border, bottom=bottom_border)
        return border

    def general_settings(self, sheet):
        sheet.column_dimensions["A"].width = self.client  # 盘口名
        sheet["A1"].font = Font(bold=True)
        sheet["A1"].border = Border()
        sheet["A1"].alignment = Alignment()
        sheet.column_dimensions["B"].width = 2
        sheet["B1"] = ""
        sheet["B1"].border = Border()
        # sheet.column_dimensions["A"].fill = PatternFill("solid", fgColor="D9D9D9")  # 此行不知道為什麼不能生效，只有沒有數據的地方套用到
        sheet.column_dimensions["C"].width = self.index  # 序號
        sheet.column_dimensions["D"].width = self.date  # 日期
        sheet.column_dimensions["E"].width = self.subject  # 憑證種類
        sheet.column_dimensions["F"].width = self.member  # 人數
        sheet.column_dimensions["G"].width = self.quantity  # 筆數
        sheet.column_dimensions["H"].width = self.amount  # 金額
        sheet.column_dimensions["I"].width = self.currency  # 幣別

    def apply_style(self):
        workbook = openpyxl.load_workbook(self.summary_path)
        for sheet_name in self.sheet_list:
            sheet = workbook[sheet_name]
            row_index = 1
            max_row = sheet.max_row
            self.general_settings(sheet)  # 到幣別之前

            if sheet_name == "單客戶":
                columns_mapping = self.one_site_columns
                site_condition = True
            else:
                columns_mapping = self.total_site_columns
                sheet.column_dimensions[columns_mapping["子客戶"]].width = self.site  # 站点
                site_condition = False
            sheet.column_dimensions[columns_mapping["憑證總額"]].width = self.certificate_amount  # 總額
            while row_index <= max_row:
                if row_index == 1:
                    self.apply_header_style(sheet, columns_mapping, site_condition)
                else:
                    self.apply_not_header_style(sheet, row_index, columns_mapping, site_condition)

                self.apply_date_style(sheet, row_index)
                self.apply_columns_number_format(sheet, row_index)

                self.apply_columns_fill(sheet, row_index, columns_mapping)
                self.apply_columns_alignment(sheet, row_index, columns_mapping, site_condition)
                self.apply_both_sides_board(sheet, row_index, site_condition)
                self.apply_hyperlinks(sheet, row_index, columns_mapping)
                self.apply_columns_font(sheet, row_index, columns_mapping)
                self.apply_function(sheet, row_index, columns_mapping)
                row_index += 1
            sheet.freeze_panes = "A2"  # 冻结首行
            sheet.auto_filter.ref = sheet.dimensions  # 自动筛选
        workbook.save(self.summary_path)

    @staticmethod
    def apply_date_style(sheet, index):
        # 日期
        column_range = "D" + str(index)
        sheet[column_range].number_format = "yyyy/mm/dd"

    def apply_header_style(self, sheet, columns_mapping, site_condition):
        if site_condition:
            cells = sheet[columns_mapping["序号"] + "1:" + columns_mapping["币别"] + "1"]
        else:
            cells = sheet[columns_mapping["序号"] + "1:" + columns_mapping["子客戶"] + "1"]
        font, fill, border, alignment = self.header()
        for row in cells:
            for cell in row:
                cell.font = font
                cell.fill = fill
                cell.border = border
                cell.alignment = alignment

    def apply_not_header_style(self, sheet, index, columns_mapping, site_condition):  # 單数
        if site_condition:
            column_range = columns_mapping["序号"] + str(index) + ":" + columns_mapping["币别"] + str(index)
        else:
            column_range = columns_mapping["序号"] + str(index) + ":" + columns_mapping["子客戶"] + str(index)
        cells = sheet[column_range]
        if (index % 2) != 0:
            font, fill, border, alignment = self.odd_row()
        else:
            font, fill, border, alignment = self.even_row()
        for row in cells:
            for cell in row:
                cell.font = font
                cell.fill = fill
                cell.border = border
                cell.alignment = alignment

    @staticmethod
    def apply_columns_fill(sheet, index, columns_mapping):
        # 盤口名稱
        cell = sheet[columns_mapping["客戶"] + str(index)]
        cell.fill = PatternFill("solid", fgColor="D9D9D9")
        # 差额
        column_range = columns_mapping["人數差額"] + str(index) + ":" + columns_mapping["金額差额"] + str(index)
        cells = sheet[column_range]
        for row in cells:
            for cell in row:
                cell.fill = PatternFill(fill_type="solid", fgColor="E2EFDA")
        # 检查
        cell = sheet[columns_mapping["檢查"] + str(index)]
        cell.fill = PatternFill(fill_type="solid", fgColor="DCE6F1")

    @staticmethod
    def apply_columns_alignment(sheet, index, columns_mapping, site_condition):
        if index != 1:
            # 三級科目
            cell = sheet[columns_mapping["憑證種類"] + str(index)]
            cell.alignment = Alignment(horizontal="distributed", vertical="center")
            # 幣別
            cell = sheet[columns_mapping["币别"] + str(index)]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if not site_condition:
                cell = sheet[columns_mapping["子客戶"] + str(index)]
                cell.alignment = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def apply_columns_number_format(sheet, index):
        if index != 1:
            # 人數 筆數
            column_range = "F" + str(index) + ":" + "G" + str(index)
            cells = sheet[column_range]
            for row in cells:
                for cell in row:
                    cell.number_format = "0_ "
            # 金額
            cell = sheet["H" + str(index)]
            cell.number_format = "#,##0.00_ "

    def apply_both_sides_board(self, sheet, index, site_condition):
        if index != 1:
            # 最左側
            cell = sheet["C" + str(index)]
            cell.border = self.far_left_border()

            # 最右側
            if site_condition:
                cell = sheet["I" + str(index)]
            else:
                cell = sheet["J" + str(index)]
            cell.border = self.far_right_border()

    @staticmethod
    def apply_hyperlinks(sheet, index, columns_mapping):
        if index != 1:
            cell = sheet[columns_mapping["路徑"] + str(index)]
            link = cell.value
            cell.hyperlink = link

    @staticmethod
    def apply_columns_font(sheet, index, columns_mapping):
        if index != 1:
            column_range = columns_mapping["路徑"] + str(index) + ":" + columns_mapping["檢查"] + str(index)
            cells = sheet[column_range]
            for row in cells:
                for cell in row:
                    cell.font = Font(name='楷体')

    @staticmethod
    def apply_function(sheet, index, columns_mapping):  # 写入公式空白必须为'""'，f字串也必须是''，单引号
        if index != 1:
            mould_people_column = columns_mapping["人数"]
            mould_count_column = columns_mapping["笔数"]
            mould_amount_column = columns_mapping["金额"]
            certificate_people_column = columns_mapping["憑證人數"]
            certificate_count_column = columns_mapping["憑證筆數"]
            certificate_amount_column = columns_mapping["憑證總額"]
            people_diff_column = columns_mapping["人數差額"]
            count_diff_column = columns_mapping["筆數差額"]
            amount_diff_column = columns_mapping["金額差额"]
            check_column = columns_mapping["檢查"]
            empty_str = '""'
            mould_column = mould_people_column
            certificate_column = certificate_people_column
            cell = sheet[people_diff_column + str(index)]
            cell.value = f'=IFERROR({mould_column}{index}-{certificate_column}{index},{empty_str})'

            mould_column = mould_count_column
            certificate_column = certificate_count_column
            cell = sheet[count_diff_column + str(index)]
            cell.value = f'=IFERROR({mould_column}{index}-{certificate_column}{index},{empty_str})'

            mould_column = mould_amount_column
            certificate_column = certificate_amount_column
            cell = sheet[amount_diff_column + str(index)]
            cell.value = f'=IFERROR({mould_column}{index}-{certificate_column}{index},{empty_str})'

            cell = sheet[check_column + str(index)]
            cell.value = (f'=IF(AND({certificate_people_column}{index}<>{empty_str},'
                          f'{certificate_count_column}{index}<>{empty_str},'
                          f'{certificate_amount_column}{index}<>{empty_str}),'
                          f'AND({people_diff_column}{index}=0,'
                          f'{count_diff_column}{index}=0,'
                          f'{amount_diff_column}{index}=0),'
                          f'FALSE)')

